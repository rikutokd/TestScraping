import requests
import openpyxl
from bs4 import BeautifulSoup

load_url = "https://news.yahoo.co.jp/categories/it"
html = requests.get(load_url)
soup = BeautifulSoup(html.content, "html.parser")

topic = soup.find(id="uamods-topics")

# Excelブック生成
wb = openpyxl.Workbook()
ws = wb["Sheet"]

# 列の幅を変える
ws.column_dimensions["A"].width = 30

topic = soup.find(id="uamods-topics")
for i, element in enumerate(topic.find_all("a")):
    title = element.text
    url = element.get("href")
    print(title)
    print(url)
        
    ws.cell(i+1,1,value = title)
    ws.cell(i+1,2).hyperlink = url

wb.save("todayYahooNews.xlsx")